from openpyxl import Workbook, load_workbook
from openpyxl.utils.exceptions import InvalidFileException


workbook_name = input('Enter file path: ')
wb = load_workbook(workbook_name)
main_ws = wb.worksheets[0]


# Gets the number of rows in an excel dataset
def get_rows():
    number = 0
    col = 'A'
    index = 1
    cell = col + str(index)

    while main_ws[cell].value is not None:
        index += 1
        number += 1
        cell = col + str(index)
    return number


# Creates a list of names
def make_list():
    rows = get_rows()
    names = []
    col = 'A'
    index = 1
    cell = col + str(index)
    for row in range(1, rows + 1):
        if not main_ws[cell].value in names:
            names.append(main_ws[cell].value)
        index += 1
        cell = col + str(index)
    return names


#Creates a dictionary that where each key word is taken from the list
def make_dictionary():
    data = {}
    name_col = 'A'
    amt_col = 'B'
    rows = get_rows()
    names = make_list()

    for name in names:
        index = 1
        amounts = []
        for row in range(1, rows + 1):

            name_cell = name_col + str(index)
            if name == main_ws[name_cell].value:
                amt_cell = amt_col + str(index)
                amounts.append(main_ws[amt_cell].value)
            index += 1
        updated_dict = {name: amounts}
        data.update(updated_dict)

    return data


def reconcile(da):

    for name in da:

        delete_index = []
        to_delete = []

        for i in range(0, len(da[name])):
            for j in range(0, len(da[name])):
                if da[name][i] + da[name][j] == 0:
                    if i not in delete_index and j not in delete_index:
                        delete_index.append(i)
                        delete_index.append(j)

        for i in range(0, len(da[name])):
            if i in delete_index:
                to_delete.append(da[name][i])

        for i in range(0, len(to_delete)):
            da[name].remove(to_delete[i])

    make_reconciled_ws(da)


def make_reconciled_ws(data_map_new):
    rec_ws = wb.create_sheet('Reconciled')
    name_col = 'A'
    value_col = 'B'
    row = 1
    name_cell = name_col + str(row)
    value_cell = value_col + str(row)
    for key in data_map_new:
        for value in data_map_new[key]:
            rec_ws[name_cell] = key
            rec_ws[value_cell] = value
            row += 1
            name_cell = name_col + str(row)
            value_cell = value_col + str(row)


data_map = make_dictionary()
reconcile(data_map)
wb.save(workbook_name)
